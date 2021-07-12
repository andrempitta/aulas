# https://www.youtube.com/watch?v=7YS6YDQKFh0&t=1790s
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pandas as pq

dadosw = [['\x01'], ['I20100'], ['27/05/21', '14:40'], [], ['Posto', 'So', 'Francisco'], ['Fortaleza-Ce'], ['J.', 'Thomaz', 'Coelho', '410'], ['Ipiranga'], [], ['INVENTRIO', 'NO', 'TANQUE'], [], ['TANQ', 'PRODUTO', 'VOLUME', 'TC-VOLUME', 'VOLUME', 'ALTURA', 'GUA', 'TEMPER'], ['1', 'Gasolina', 'Comum', '26124', '25788', '4703', '2013.97', '0.00', '30.30'], ['2', 'Gasolina', 'Aditivada', '21464', '21172', '8562', '1713.46', '0.00', '30.79'], ['3', 'Gasolina', 'Comum', '25646', '25310', '5554', '1952.66', '0.00', '30.48'], ['4', 'Etanol', '13897', '13717', '1337', '2184.44', '0.00', '31.34'], ['5', 'Diesel', 'Comum', '4362', '4321', '10758', '842.63', '0.00', '31.49'], ['6', 'Gasolina', 'Comum', '14054', '13878', '17146', '1175.20', '0.00', '30.04'], [], ['\x03']]

dados = (20100, 410,26124, 25788, 4703, 2013.97, 0.00, 30.30, 2,21464, 21172, 8562, 1713.46, 0.00, 30.79, 3, 25646, 25310, 5554, 1952.66, 0.00, 30.48, 4, 13897, 13717, 1337, 2184.44, 0.00, 31.34, 5, 4362, 4321, 10758, 842.63, 0.00, 31.49, 6, 14054, 13878, 17146, 1175.20, 0.00, 30.04)

a = pd.DataFrame()

wb = Workbook()
ws = wb.active
ws.title = 'soma_ultima_celular'
for dado in range(0, len(dados) - 1):
    ws.append(dados[dado])



wb.save('excel_teste_2')



# para transferir de um sheet (planilha) para outro!
# wb = openpyxl.load_workbook('file1.xlsx')
# wb1 = openpyxl.load_workbook('file2.xlsx')
# sheet = wb["R"]
# sheet1 = wb1["Rt"]
# for i in range(1, 100):
#     for j in range(1, 26):
#         sheet1.cell(row=i,column=j).value = sheet.cell(row=i,column=j).value
# wb1.save('file2.xlsx')


"""

data = {
    "André" : {
        "matemática" : 9.5,
        "ciências" : 9.0,
        "inglês" : 7.5,
        "geografia" : 8.2
    },
    "Ludmila" : {
        "matemática" : 9.2,
        "ciências" : 9.4,
        "inglês" : 9.5,
        "geografia" : 7.2
    },
    "Marcela" : {
        "matemática" : 9.5,
        "ciências" : 9.9,
        "inglês" : 9.5,
        "geografia" : 8.2
    },
    "Leonardo" : {
        "matemática" : 10,
        "ciências" : 9.0,
        "inglês" : 8.5,
        "geografia" : 9.2
    }
}

wb = Workbook()
ws = wb.active
ws.title = "Notas"

cabecalho = ["Nomes"] + list(data['André'].keys())
ws.append(cabecalho)

# insere os dados na planilha
for pessoa in data:
    notas = list(data[pessoa].values())
    ws.append([pessoa] + notas)

# insere formula de média
for col in range(2, len(data['André']) + 2):
    char = get_column_letter(col)
    ws[char + '6'] = f"=SUM({char + '2'}:{char + '5'})/{len(data)}"

# cor no cabeçalho
for col in range(1,6):
    ws[get_column_letter(col) + '1'].font = Font(bold=True, color="0099CCFF")

wb.save("Notas_teste.xlsx")
"""




# # Mover células
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# ws.move_range('c1:d11', rows=2, cols=-1)
# wb.save('Teste_Andre.xlsx')


# # Deletar linha e coluna vazia
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# # ws.delete_rows(4)
# ws.delete_cols(2)
# wb.save('Teste_Andre.xlsx')


# # Inserir linha e coluna vazia
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# # ws.insert_rows(4)
# ws.insert_cols(2)
# wb.save('Teste_Andre.xlsx')


# UnMarge celulas
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# ws.unmerge_cells('A1:D1')
# wb.save('Teste_Andre.xlsx')


# # Marge celulas
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# ws.merge_cells('A1:D1')
# wb.save('Teste_Andre.xlsx')


# # Altera o valor das células nesse range para a coordenada dela 'A1, A2, etc...)
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# for linha in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         ws[char + str(linha)] = char + str(linha)
# wb.save('Teste_Andre.xlsx')


# print(chr(65)) # caracter A
# print(chr(66)) # caracter B

# # Imprime o valor das células nesse range
# wb = load_workbook('Teste_Andre.xlsx')
# ws = wb.active
# for linha in range(1, 11):
#     for col in range(1, 5):
#         char = get_column_letter(col)
#         print(ws[char + str(linha)].value)
# wb.save('Teste_Andre.xlsx')



# wb = Workbook()
# ws = wb.active
# ws.title = 'Teste'
# ws.append(['André', 'Macedo', 'Pitta', '!'])
# ws.append(['André', 'Macedo', 'Pitta', '!'])
# ws.append(['André', 'Macedo', 'Pitta', '!'])
# wb.save('Teste_Andre.xlsx')


# wb = load_workbook('Notas_teste2.xlsx')
# ws = wb.active # ws = wb.['Sheet1']  #  para trabalhar na pasta Sheet1
# wb.create_sheet('Teste') # cria nova pasta
# print(wb.sheetnames) #  mostras as pastas
# wb.save("Notas_teste2.xlsx")


# print(ws.title)  # nome das pasta
# print(ws['A1'].value)
# print(ws['b3'].value)
# ws['E3'].value = 123
# ws['E4'].value = "Teste"
Downloads


/html/body/form/table/tbody/tr[4]/td/table/tbody/tr[2]/td/div[2]/div[1]/div[1]/a

<input name="pass" id="pass" type="password" class="in col-10 ui-keyboard-input ui-widget-content ui-corner-all ui-keyboard-lockedinput" aria-haspopup="true" role="textbox" readonly="readonly">

